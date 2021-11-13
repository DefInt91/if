#!/usr/bin/env python
# coding: utf-8

# ![ITOC](DFT.png)
# # PagerDuty Alerts - Analysis & Statstics v3.0

# ###版更內容###
# 2021/01/10 註記improt openpyxl 修改無法跑出檔案的問題 (未使用到的套件)--V3
# 2021/08/29 新增兩個列(columns) "Host" & "Alert_type", 新增Chirs進篩選名單--V4
# 2021/09/15 新增Jiajin進篩選名單--V4.1

# ##  Summary 
# 1. Create folder to save output files
# 2. Import .csv files
# 3. Filter out information about ITOC
# 4. Data processing
#    1. Alert Flood (?)
#    2. Shift per Day(Weekday) + plt.pie
#    3. Alerts per Weekday
#    4. Total alerts per day + Top 10 alerts
#    5. Alerts resolved by ITOC
#    6. Escalated to DFT/SMT
#    7. Alerted in work/after hours * shift + plt.pit
# 7. Use plot and bar graph to show date
# 7.5 Add two columns named Host & Alert_type
# 8. Output data to ~

# ### DataFrames have been named
#  1. alerts_ITOC <- report[report.service_id == 'PG1SJTA']
#  2. alerts <- pd.DataFrame(alerts_ITOC['description'].value_counts().head(10))
#  3. alerts_counter <- alerts_ITOC['description'].str.split(':', 1, True)
#  4. host <- pd.DataFrame(alerts_counter['host'].value_counts().head(10))
#  5. issues <- pd.DataFrame(alerts_counter['issue'].value_counts().head(10))
#  6. alerts_date <- alerts_ITOC['created_on'].str.split('T', 1, True)
#  7. weekday <- pd.DataFrame(alerts_date['dayOfWeek'].value_counts())
#  8. hourOfDay <- alerts_date['HH'].value_counts()
#  9. shift_count_per_date <- alerts_date['shifts'] + alerts_date['new_date']
#  10. shift_count_per_hour <- alerts_date['shifts'] + alerts_ITOC['created_on']
#  11. urgency <- pd.DataFrame(alerts_ITOC['urgency'].value_counts()) 

# ### Graphs go line 425
# 1. shifts_pie <- alerts_date['shifts'].value_counts()
# 2. alerts_per_weekday <- weekday['dayOfWeek'], weekday['count']
# 3. alerts_per_hour <- hourOfDay['24-hour'], hourOfDay['count']
# 4. add 3 chart in [Shift_stats],[Comparison] - 2020/12/01


# ### Function
# * 1. shift_of_day
#     * Day shift 07 to 15
#     * Late shift 15 to 23
#     * Night shift 23 to 07
# * 2. Alert flood
#     * Received over 100 alerts in 1 hour
# * 3. Escalation to DFT/SMT in work/after hours

# In[323]:


#引入模組
from contextlib import asynccontextmanager
from typing import Type
from numpy.lib.npyio import load
from openpyxl import load_workbook
import pandas as pd
# import matplotlib.pyplot as plt
import os
from pandas.core.frame import DataFrame
from pandas.io.formats import style
import pytz
import numpy as np
import datetime
import time
import calendar
import xlsxwriter
import requests
# import json
import csv
from datetime import date
import openpyxl




# ## 1. Create folder to save output files

# In[324]:

########################################################################################
### Set up a folder on Desktop
def GetDesktopPath():
    return os.path.join(os.path.expanduser("~"), 'Desktop')

def mkdir(path):
    #判斷目錄是否存在
    #存在：True
    #不存在：False
    folder = os.path.exists(path)

    #判斷結果
    if not folder:
        #如果不存在，則建立新目錄
        os.makedirs(path)
        print('-----建立成功-----')

    else:
        #如果目錄已存在，則不建立，提示目錄已存在
        print(path+'目錄已存在')
path = GetDesktopPath() + '\\alert_stat'
mkdir(path)
########################################################################################

### Get the correct date(Since/Until)
aaa, last_Mon, last_Sun = datetime.date.today(), datetime.date.today(), datetime.date.today()
one_day = datetime.timedelta(days=1)
seven_day = datetime.timedelta(days=7)
# print(str(last_Mon.month) + '-' + str(last_Mon.day) + '_' + str(last_Mon.weekday()))

last_month = date.today().month-1
month = date.today().month

xday=datetime.datetime.now() #取出現在時間日期 參考http://blog.alarmchang.com/?p=230
xday.day#取得今日日期
xday.month


# 取得前一週日期 週一到週日
if last_Mon.weekday() == 0:
    last_Mon -= seven_day
    last_Sun -= one_day
    # print(last_Mon,last_Mon.weekday(),last_Sun,last_Sun.weekday())
else:
    while last_Mon.weekday() != 0: # 如果不等於1則減一天, 星期一 = 0,  星期二 = 1, 星期三 = 2,  星期四 = 3,  星期五 = 4,  星期六 = 5,  星期日 = 6
        last_Mon -= one_day
        last_Sun -= one_day
        # print(last_Mon,last_Mon.weekday())
    last_Mon -= seven_day
    last_Sun -= one_day
    # print("Get Time range: " + str(last_Mon) + '_' + str(last_Mon.weekday()) + '_' + str(last_Sun) + '_' + str(last_Sun.weekday()))
print("Time range: " + str(last_Mon) + ' to ' + str(last_Sun))


### Date format: 'yyyy-MM-ddThh:mm:ss'
if last_Mon.month < 10:
        if last_Mon.day < 10:
            since = str(last_Mon.year)+'-0' + str(last_Mon.month) +'-0'+ str(last_Mon.day) + 'T00:00:00'
            # until = str(last_Sun.year)+'-0' + str(last_Sun.month) +'-0'+ str(last_Sun.day) + 'T23:59:59'
            # print (since,'thu.month < 10, thu.day < 10')
        else:
            since = str(last_Mon.year)+'-0' + str(last_Mon.month) +'-'+ str(last_Mon.day) + 'T00:00:00'
            # print (since,'thu.month < 10, thu.day > 10')
elif (last_Mon.month == 10): 
        if last_Mon.day < 10:
            since = str(last_Mon.year)+'-' + str(last_Mon.month) +'-0'+ str(last_Mon.day) + 'T00:00:00'
            # until = str(last_Sun.year)+'-' + str(last_Sun.month) +'-'+ str(last_Sun.day) + 'T23:59:59'
            # print (since,'thu.month = 10, thu.day < 10')
        else:
            since = str(last_Mon.year)+'-' + str(last_Mon.month) +'-'+ str(last_Mon.day) + 'T00:00:00'
            # print (since,'thu.month = 10, thu.day > 10')
else:
        if last_Mon.day <10:
            since = str(last_Mon.year)+'-' + str(last_Mon.month) +'-0'+ str(last_Mon.day) + 'T00:00:00'
            # until = str(last_Sun.year)+'-' + str(last_Sun.month) +'-'+ str(last_Sun.day) + 'T23:59:59'
            # print (since,'thu.month > 10, thu.day < 10')
        else:
            since = str(last_Mon.year)+'-' + str(last_Mon.month) +'-'+ str(last_Mon.day) + 'T00:00:00'
            # print (since,'thu.month > 10, thu.day > 10')


if last_Sun.month < 10:
        if last_Sun.day < 10:
            until = str(last_Sun.year)+'-0' + str(last_Sun.month) +'-0'+ str(last_Sun.day) + 'T23:59:59'
            # print (since,'wed.month < 10, wed.day < 10')
        else:
            until = str(last_Sun.year)+'-0' + str(last_Sun.month) +'-'+ str(last_Sun.day) + 'T23:59:59'
            # print (since,'wed.month < 10, wed.day > 10')
elif (month == 10): 
        if last_Sun.day < 10:
            until = str(last_Sun.year)+'-' + str(last_Sun.month) +'-0'+ str(last_Sun.day) + 'T23:59:59'
            # print (since,'wed.month = 10, wed.day < 10')
        else:
            until = str(last_Sun.year)+'-' + str(last_Sun.month) +'-'+ str(last_Sun.day) + 'T23:59:59'
            # print (since,'wed.month = 10, wed.day > 10')
else:
        if last_Sun.day < 10:
            until = str(last_Sun.year)+'-' + str(last_Sun.month) +'-0'+ str(last_Sun.day) + 'T23:59:59'
            # print (since,'wed.month > 10, wed.day < 10')
        else:
            until = str(last_Sun.year)+'-' + str(last_Sun.month) +'-'+ str(last_Sun.day) + 'T23:59:59'
            # print (since,'wed.month > 10, wed.day > 10')

# print(since,until)
### Call PD API to get report for specific time(Since/Until)
url = 'https://bigasia.pagerduty.com/api/v1/reports/raw/incidents.csv?filters[urgency]=high%2Clow'

######################################################
############# Get your API token from PD #############
######################################################

api_key = '18hLb33RdQax-C9DasgR'

######################################################
######################################################

headers = {
    'Accept': 'application/vnd.pagerduty+json;version=2',
    'Authorization': 'Token token=' + api_key
}
payload = {
    'since': since,
    'until': until,
    'rollup':'weekly',
    'time_zone': "Asia/Taipei",
}
r = requests.get(url, headers=headers, params=payload)

incidents = csv.reader(r.text.splitlines(), delimiter=',')
my_header = list(incidents)

### Write PD report(raw data) to csvfile
with open(path + '\\weekly_report_' + str(date.today()) + '.csv','w',newline='') as csvfile:
    writer = csv.writer(csvfile, my_header[0])
    # writer.writeheader()
    for i in my_header:
        # print(i)
        writer.writerow(i)


# # 2. Import .csv files
# 1. 欄位
#     1. incident_number
#     2. description
#     3. service_id
#     4. escalation_policy_name
#     5. created_on
#     6. assignment_count
#     7. assigned_to_user_names
#     8. urgency
#     9. resolved_by_user_name

# In[325]:


report = pd.read_csv(os.path.join(os.path.expanduser("~"), 'Desktop') + '\\alert_stat' + '\\weekly_report_' + str(date.today()) +'.csv',usecols=['incident_number', 'description', 'service_id', 'escalation_policy_name', 'created_on', 'assignment_count', 'acknowledged_by_user_names', 'seconds_to_resolve', 'urgency', 'resolved_by_user_name'])
# report = pd.read_csv(os.path.join(os.path.expanduser("~"), 'Desktop') + '\\alert_stat' + '\\testtestestestestst' +'.csv',usecols=['incident_number', 'description', 'service_id', 'escalation_policy_name', 'created_on', 'assignment_count', 'acknowledged_by_user_names', 'seconds_to_resolve', 'urgency', 'resolved_by_user_name'])


# In[326]:


report.head()


# # 3. Filter out information about ITOC

# In[327]:


alerts_ITOC = report[report.service_id == 'PD806Y5']
alerts_ITOC = alerts_ITOC.reset_index(drop=True)


# In[328]:


### total amount of ITOC's alert this month

year = datetime.datetime.now().year
last_month = datetime.datetime.now().month -1

if last_month== 0:
    last_month=12

month = {
    'January': 1,
    'February': 2,
    'March': 3,
    'April': 4,
    'May': 5,
    'June': 6,
    'July': 7,
    'August': 8,
    'September': 9,
    'October': 10,
    'November': 11,
    'December': 12
}

month_short = {
    'Jan': 1,
    'Feb': 2,
    'Mar': 3,
    'Apr': 4,
    'May': 5,
    'Jun': 6,
    'Jul': 7,
    'Aug': 8,
    'Sep': 9,
    'Oct': 10,
    'Nov': 11,
    'Dec': 12
}


for m,v in month.items():
    if v == last_month:
        last_month_w = m

#取得月份短名稱
for m,v in month_short.items():
    if v == last_Mon.month:
        last_Mon_w = m

for m,v in month_short.items():
    if v == last_Sun.month:
        last_Sun_w = m
        
#alerts_ITOC = alerts_ITOC.reset_index(drop=True)

if last_Mon.year == year:
    print("Total amount of ITOC's alerts for " + str(last_Mon_w) + '_' + str(last_Mon.day) +'-' +str(last_Sun_w) + '_' + str(last_Sun.day) + ": " + str(len(alerts_ITOC)))
else:
    print("Total amount of ITOC's alerts for " + str(last_Mon.year)+' '+str(last_Mon_w) + '_' + str(last_Mon.day) +'-'+str(year) + ' ' +str(last_Sun_w) + '_' + str(last_Sun.day) + ": " + str(len(alerts_ITOC)))
print("Loading...")

countx=len(alerts_ITOC)


# # 4. Data processing
# 1. Split "date" and "time"
# 2. Alert Flood
# 3. Shift per Day(Weekday) + plt.pie
# 4. Alerts per Weekday
# 5. Total alerts per day + Top 10 alerts
# 6. Alerts resolved by ITOC
# 7. Escalated to DFT/SMT
# 8. Alerted in work/after hours * shift + plt.pit


# ## 1. Split "date" and "time"
# In[329]:


### Split 'created_on' to 'YYYY', 'MM', 'DD', 'HH', 'mm', 'SS'


alerts_date = alerts_ITOC['created_on'].str.split('T', 1, True)
alerts_date.columns = ['date', 1]

alerts_time = alerts_date[1].str.split('+', 1, True)
alerts_time.columns = ['time', 1]
alerts_time = alerts_time['time'].str.split(':', 2, True)
alerts_time.columns = ['HH', 'mm', 'SS']

alerts_date = alerts_date['date'].str.split('-', 2, True)
alerts_date.columns = ['YYYY', 'MM', 'DD']

### Join two dataframe - alerts_date and alerts_time
alerts_date = alerts_date.join(alerts_time)
alerts_date = alerts_date.reset_index(drop=True)

alerts_ITOC['Hour_Of_Day'] = alerts_time['HH']

#alerts_date.dtypes
#alerts_date


# In[330]:


## Get the current 'year', and the correct 'last month'
current_month = time.strftime("%m", time.localtime())

#current_time = datetime.datetime.now().timetuple()
#last_month = current_time.tm_mon -1



#
if last_month == 0:
    last_month = 12
    year -= 1
    
#
#if datetime.datetime.now().month:
#    if last_month <10:
#        start_date = (str(year)+"-0"+str(last_month)+"-01")
#        end_date = (str(year)+"-0"+str(datetime.datetime.now().month)+"-01")
#    else:
#        start_date = (str(year)+"-"+str(last_month)+"-01")
#        end_date = (str(year)+"-"+str(datetime.datetime.now().month)+"-01")


# ## 2. Alert Flood

# In[331]:


alerts_ITOC['created_on_dt'] = pd.to_datetime(alerts_ITOC.created_on)
alerts_ITOC['created_on_weekday'] = pd.to_datetime(alerts_ITOC['created_on_dt']).dt.strftime('%m/%d/%Y %A')


Alert_Flood = alerts_ITOC['created_on_weekday'].value_counts()
Alert_Flood[Alert_Flood > 100]


# ## 3. Shift per Day(Weekday) + plt.pie

# In[332]:


### Set shifts for the corresponding time
def shift_of_day(df, col_name):
    #shift_of_day={
    #    "Day":range(7,15),
    #    "Late":range(15,23),
    #    "Night":range(23,7)
    #}
    
    counter = 0
    
    col = df[col_name].astype(int) 
    
    for h in col:        
        if h >= 7 and h < 15:
            col[counter] = 'Day'
            #print(str(h) + '= day')
        elif h >= 15 and h < 23:
            col[counter] = 'Late'
            #print(str(h) + '= late')
        elif h == 23 or h < 7:    
            col[counter] = 'Night'
        else:
            print(str(h) + 'is invalid value')
            break
        counter += 1 
    return col

alerts_ITOC['shifts'] = shift_of_day(alerts_date,'HH')





# In[333]:


### Get weekday in value and word
#alerts_ITOC['dayOfWeek'] = alerts_ITOC['date'].map(lambda x: pd.to_datetime(str(x)).dayofweek+1)
#alerts_ITOC['weekday'] = alerts_ITOC['date'].map(lambda x: pd.to_datetime(str(x)).day_name())

#alerts_ITOC['created_on_dt'] = pd.to_datetime(alerts_ITOC.created_on)

#localtime = alerts_ITOC.created_on_dt

#ts = pd.Series(np.random.randn(len(localtime)),index=localtime)
#ts.head(5)

#alert_local_time = ts.tz_convert('Asia/Shanghai')
##alerttime.head()
#
#alerts_ITOC['created_on_dt'] = alert_local_time.index
#alerts_ITOC['created_on'] = alert_local_time.index
alerts_ITOC['created_on'] = alerts_ITOC['created_on'].astype(str)

alerts_ITOC['date'] = [x.date() for x in alerts_ITOC['created_on_dt']]
alerts_ITOC['date_str'] = alerts_ITOC['date'].astype(str)

alerts_ITOC['time'] = [x.time() for x in alerts_ITOC['created_on_dt']]
alerts_ITOC['time_str'] = alerts_ITOC['time'].astype(str)


alerts_ITOC['dayOfWeek'] = pd.DatetimeIndex(alerts_ITOC['date']).dayofweek+1
alerts_ITOC['weekday'] = pd.DatetimeIndex(alerts_ITOC['date']).day_name()

#alerts_ITOC['dayOfWeek'].value_counts()


# In[334]:


#created_on_weekday = pd.DataFrame()
#counter = 0
#created_on_weekday

#alerts_ITOC['created_on_weekday'] = pd.to_datetime(alerts_ITOC['created_on_dt']).dt.strftime('%m/%d/%Y %A')
#for d in alerts_ITOC.created_on_dt:
#    #d = pd.to_datetime(str(d))
#    #wd = pd.to_datetime(str(d)).day_name()
#    #d_str = datetime.date.strftime(d, "%m/%d/%Y %A")
#    #created_on_weekday[counter] = str(d_str) + ' ' + d.day_name()
#    created_on_weekday[counter] = datetime.date.strftime(d, "%m/%d/%Y %A")
#    counter += 1
#
#alerts_ITOC['created_on_weekday'] = created_on_weekday


#shift_count_per_hour = shift_count_per_hour.pivot_table(index='created_on_weekday', columns='shifts', aggfunc='count')
#shift_count_per_hour


# In[335]:


alert_count_per_shift = pd.DataFrame()
alert_count_per_shift['weekday'] = alerts_ITOC['created_on_weekday']
alert_count_per_shift['shifts'] = alerts_ITOC['shifts']
alert_count_per_shift['value'] = 1

alerts_ITOC['created_on_weekday'].sort_index(ascending=False)
alert_count_per_shift_pt = alert_count_per_shift.pivot_table(index=alert_count_per_shift.weekday, columns='shifts', fill_value=0, values='value', aggfunc='count')
#alert_count_per_shift_pt = alert_count_per_shift.pivot_table(index=alert_count_per_shift.weekday, columns='shifts', aggfunc='count')


#shifts = ["Day", "Late", "Night"]

#alert_count_per_shift = alert_count_per_shift.reindex(shifts,level='shifts', axis=1)
#alert_count_per_shift
#alert_count_per_shift['Total'] = alerts_ITOC.created_on_weekday.value_counts()
alert_count_per_shift_pt['Total'] = alerts_ITOC.created_on_weekday.value_counts()

#alert_count_per_shift_pt


# In[336]:


shift_comparison = pd.DataFrame(alerts_ITOC['shifts'].value_counts())

shift_comparison = shift_comparison.rename_axis(last_month_w)
shift_comparison.columns = ['Count']

shift_comparison = shift_comparison.T
shift_comparison['Total'] = len(alerts_ITOC)
shift_comparison


# In[337]:


### plt.pie

# shifts_pie = alerts_ITOC['shifts'].value_counts()
# size = list(shifts_pie.values.astype(int))
# labels = tuple(shifts_pie.index.astype(str))
# plt.pie(size, labels = labels, autopct='%1.1f%%')
# plt.axis('equal')
# plt.savefig(path + '\\shifts_pie.png')




# In[338]:


alerts_ITOC.head(1)


# ## 4. Alerts per Weekday

# In[339]:


### 3.Alerts per Weekday
alert_count_per_weekday = pd.DataFrame()
alert_count_per_weekday['created_on_weekday'] = alerts_ITOC['created_on_weekday']
alert_count_per_weekday['weekday'] = alerts_ITOC['weekday']
alert_count_per_weekday['value'] = 1

#from pandas.api.types import CategoricalDtype
#
#

#cat_type = CategoricalDtype(categories=cats, ordered=True)
#alert_count_per_weekday['weekday'] = alert_count_per_weekday['weekday'].astype(cat_type)
##alert_count_per_weekday.dtypes
#alert_count_per_weekday

#alert_count_per_weekday['weekday'] = alert_count_per_weekday['weekday'].astype('category', categories=cats, ordered=True)
#alert_count_per_weekday['weekday']

#alert_count_per_weekday['weekday'] = pd.Categorical(alert_count_per_weekday['weekday'], categories=['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday', 'Sunday'],ordered=True)


alert_count_per_weekday_pt = alert_count_per_weekday.pivot_table(index=alert_count_per_weekday.created_on_weekday, columns='weekday', fill_value='', values='value', aggfunc='count')
cats = [ 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
alert_count_per_weekday_pt = alert_count_per_weekday_pt.reindex(cats, axis=1)
#alert_count_per_weekday_pt = pd.pivot_table(alert_count_per_weekday, index='created_on_weekday', columns=['weekday'], aggfunc='count')



#alert_count_per_weekday_pt = alert_count_per_weekday.groupby(['weekday']).count().reindex(cats) 

#alert_count_per_weekday.weekday = alert_count_per_weekday.weekday.date.strftime('%A')


#column_order = ["Monday", "Thursday", "Wednesday", "Tuesday", "Friday", "Saturday", "Sunday"]
#alert_count_per_weekday.reindex(column_order, axis=1)

#weekday_count = alerts_ITOC.weekday.value_counts()
##weekday_count
alert_count_per_weekday_pt['Total'] = alerts_ITOC.created_on_weekday.value_counts()
#
##alert_count_per_weekday.dropna()  
#alert_count_per_weekday = alert_count_per_weekday[['Monday', 'Thursday', 'Wednesday', 'Tuesday', 'Friday', 'Saturday', 'Sunday']].copy()

#alert_count_per_weekday.sort_values(by=2012, ascending=False).index)


# In[340]:


alert_count_per_weekday_pt


# In[341]:


### 4.Total alerts per day + Top 10 alerts

alerts_per_day = alert_count_per_weekday.groupby('created_on_weekday')[['value']].count()
alerts_per_day.columns = ['count']

#alerts_per_day = pd.DataFrame(alerts_ITOC['created_on_weekday'].value_counts())
#alerts_per_day = alerts_per_day.rename_axis('date')
#alerts_per_day.columns = ['count']
#alerts_per_day = alerts_per_day.sort_values(by='date', ascending=True).reset_index()
alerts_per_day


# In[342]:


## ii.Top 10 alerts
top_alerts = pd.DataFrame(alerts_ITOC['description'].value_counts().head(10))

top_alerts = top_alerts.rename_axis('alerts')
top_alerts.columns = ['count']
top_alerts


# In[343]:


## ii.Top host & issue
alerts_des = alerts_ITOC['description'].str.split(': ', 1, True)
alerts_des.columns = ['host', 'issue']

top_host = pd.DataFrame(alerts_des['host'].value_counts().head(10))
top_host = top_host.rename_axis('host')
top_host.columns = ['count']

top_issue = pd.DataFrame(alerts_des['issue'].value_counts().head(10))
top_issue = top_issue.rename_axis('issue')
top_issue.columns = ['count']









# In[344]:


### Count the alerts per day
#alerts_per_day = pd.DataFrame(alerts_ITOC['date'].value_counts())
#alerts_per_day = alerts_per_day.rename_axis('date')
#alerts_per_day.columns = ['count']
#alerts_per_day = alerts_per_day.sort_values(by='date',ascending=True)
#alerts_per_day


# ## 6. Escalated to DFT/SMT

# In[345]:


### Check if the alerts were assgined to 'ITOC', 'DFT', 'SMT'

def reassigned_to():
    #escalated_to={
    #    "ITOC":(assignment_count < 3) and (escalated_to == 'ITOC'),
    #    "DFT":(assignment_count >= 3) and (escalated_to == 'DFT'),
    #    "SMT":(assignment_count >= 3) and (escalated_to == 'SMT')
    #}
    
    escalated = alerts_ITOC['escalated'].astype(int) 
    assigned_to = alerts_ITOC['escalated'].astype(str)
    
    counter = 0
    
    for v in escalated:        
        if v == 0:
            assigned_to[counter] = 'ITOC'
        if v == 1:    
            assigned_to[counter] = 'DFT'
        if v == 2:   
            assigned_to[counter] = 'SMT'
        counter += 1 
        
    
    counter = 0
    
    ITOC_member = (
        'Ray Duan', 'Jackie Xu', 'Joy Huang', 'Rong Liu', 'Jin Yang', 'Peter Tsai', 'Daniel Lin', 'Anderson Liao', 'Acise Lee','Acs', 'Da Wu', 'Joseph Cheng', 'Jake Wu','Chris Lia','Eric Tzong', 'ITOC (East)', 'Jiajin Kang'
    )
    
    
    #assign_count = alerts_ITOC['assignment_count'].astype(int) 
    
    user = alerts_ITOC['resolved_by_user_name'].astype(str)
    for user_name in user:
        a = 0
        if assigned_to[counter] == 'ITOC':
            a = 1
            for name in ITOC_member:
                if user[counter] == name:
                    a += 1
        if a == 1:
            assigned_to[counter] = user[counter]
        counter += 1
    

    alerts_ITOC['assigned_to'] = assigned_to

def check_escalation_policy():
    ### 0 = ITOC, 1 = DFT, 2 = SMT
    
    counter = 0
    
    SMT = alerts_ITOC['escalation_policy_name'].str.contains('Infrastructure|DCS', na=False)
    
    col = alerts_ITOC['escalated'].astype(int)
    
    for team in SMT:
        if team == True:
            col[counter] = 2
        counter += 1
    
    ITOC = alerts_ITOC['escalation_policy_name'].str.contains('Centre', na=False)
    
    counter = 0
    for team in ITOC:
        if team == True:
            col[counter] = 0
        counter += 1

    alerts_ITOC['escalated'] = col
    reassigned_to()


# In[346]:


alerts_ITOC['escalated'] = 1
alerts_ITOC['assigned_to'] = '1'

check_escalation_policy()


# In[347]:


alerts_ITOC.assigned_to.value_counts()


# In[348]:


### Alert count per team

#a = alerts_ITOC[alerts_ITOC['escalation_policy_name'].str.contains('Infrastructure|DCS', na=False)]
#a.shape
#alerts_ITOC['escalated'].value_counts()

reassigned_to = pd.DataFrame(alerts_ITOC.escalation_policy_name.value_counts())
reassigned_to = reassigned_to.rename_axis('escalation_policy')
reassigned_to.columns = ['count']
#reassigned_to = reassigned_to.sort_values(by='escalation_policy',ascending=True)
#reassigned_to


# ## 7. Alerts resolved by ITOC

# In[349]:


### 5.Alerts resolved by ITOC
resolved_by_ITOC = alerts_ITOC[alerts_ITOC['assigned_to'].str.contains('ITOC', na=False)]
resolved_by_ITOC = resolved_by_ITOC.reset_index(drop=True)
#resolved_by_ITOC


# ## 8. Alerted in work/after hours * shift + plt.pit

# In[350]:


### 8.Alerted in work/after hours * shift + plt.pit

### Check when the alerts were triggered, it's during office hours or not 

#def check_holiday():
    

def office_horus(df, col_name):
    #office_horus={
    #    "working-hours":range(9,18),
    #    "off-hours":range(15,23),
    #    "Night":range(0,7)
    #}
    
    counter = 0
    
    
    col = df[col_name].astype(int) 
    col1 = alerts_ITOC['dayOfWeek'].astype(int)
    
    for h in col:   
        if (h >= 9 and h < 18) and (col1[counter] <= 5):
            col[counter] = 'Y'            
            #print(str(h) + '= day')
        #elif (h >= 18 or h < 9) or (col1[counter] > 5):    
        #    col[counter] = 'off-hours'
        else:
            col[counter] = 'N'
            #print(str(h) + 'is invalid value')
            #break
        counter += 1 
        
    #check_holiday()
    return col

alerts_ITOC['work_hr'] = office_horus(alerts_date,'HH') 
    
#alerts_ITOC.work_hr.value_counts()  


# In[351]:


alerts_ITOC['work_hr'].value_counts() 


# # 7. Use plot and bar graph to show date

# # 7.5 Add two columns named Host & Alert_type
alerts_host_type=pd.DataFrame(alerts_ITOC['description'].str.split(': ', 1).tolist(),columns=['Host','Alert_type'])
alerts_host_type.columns=['Host','Alert_type']
alerts_host_type['Alert_type']=alerts_host_type['Alert_type'].fillna(alerts_host_type['Host']) # refer to this URL:https://stackoverflow.com/questions/46676134/what-is-the-difference-between-combine-first-and-fillna
alerts_ITOC['Host']=alerts_host_type['Host'].astype(str)
alerts_ITOC['Alert_type']=alerts_host_type['Alert_type'].astype(str)

# # 8. Output data to ~

# In[352]:


alerts_ITOC.dtypes


# In[353]:
# bytime=pd.DataFrame()
# bytime['Hour_Of_Day']=alerts_ITOC['Hour_Of_Day']
# bytime['Date']=alerts_ITOC['date']
# bytime['value'] = 1

# bytime = bytime.pivot_table(index=bytime.Hour_Of_Day, columns='Date', fill_value=int('0'), values='value', aggfunc='count')
# def color_red(count):

#     color = 'red' if count >=13 else 'black'
#     return 'color: %s' % color

# # print(values)
# bytime.style.applymap(color_red,subset=['Date'])

# bytime['Total'] = alerts_ITOC.Hour_Of_Day.value_counts()


# countalert=pd.DataFrame()
# countalert['Alert_type']=alerts_ITOC['Alert_type']
# countalert['Date']=alerts_ITOC['date']
# countalert['value'] = 1

# # rslt_CA=countalert.loc[countalert['Date']>13]
# # print(rslt_CA)
# countalert['Total'] = alerts_ITOC.Alert_type.value_counts()

# # countalert = countalert.pivot_table(countalert[countalert.Total>13],index=countalert.Alert_type, columns='Date', fill_value='', values='value', aggfunc='count')
# countalert = countalert.pivot_table(countalert,index=countalert.Alert_type, columns='Date', fill_value='', values='value', aggfunc={'Date':'count'})

alerts_ITOC.head(1)



# In[362]:


alerts_escalation = alerts_ITOC[['escalation_policy_name', 'incident_number', 'description', 'Hour_Of_Day', 'escalated', 'work_hr', 'urgency', 'seconds_to_resolve', 'acknowledged_by_user_names', 'resolved_by_user_name' ]]
alerts_escalation.index = alerts_ITOC['created_on_weekday']


# In[355]:


### Select output column

output = alerts_ITOC[['incident_number', 'description', 'Host', 'Alert_type' ,'escalation_policy_name', 'assignment_count', 'created_on_weekday', 'date', 'time', 'Hour_Of_Day', 'weekday', 'shifts', 'assigned_to', 'acknowledged_by_user_names', 'escalated', 'work_hr', 'urgency']]

output = output.set_index(['created_on_weekday'])





# In[356]:


### Replace the index of alerts_ITOC with 'timestamp'
#alerts_ITOC.index = alerts_index['timestamp']


# In[363]:

# 檔名依上週一至本週日命名
if last_Mon.year == year:
    writer = pd.ExcelWriter(path + '\\' + str(year) + ' ' + str(last_Mon_w) + '_' + str(last_Mon.day) +'-' +str(last_Sun_w) + '_' + str(last_Sun.day) + ' - PD_Stats_Report.xlsx', engine='xlsxwriter', options={'remove_timezone': True})
    print('File Name: '+str(year) + ' ' + str(last_Mon_w) + '_' + str(last_Mon.day) +'-' +str(last_Sun_w) + '_' + str(last_Sun.day) + ' - PD_Stats_Report.xlsx')
    FN=str(year) + ' ' + str(last_Mon_w) + '_' + str(last_Mon.day) +'-' +str(last_Sun_w) + '_' + str(last_Sun.day) + ' - PD_Stats_Report.xlsx'
else:     
    writer = pd.ExcelWriter(path + '\\' + str(last_Mon.year) + ' ' + str(last_Mon_w) + '_' + str(last_Mon.day) + '-' + str(year) + ' ' +str(last_Sun_w) + '_' + str(last_Sun.day) + ' - PD_Stats_Report.xlsx', engine='xlsxwriter', options={'remove_timezone': True})
    print('File Name: '+str(last_Mon.year) + ' ' + str(last_Mon_w) + '_' + str(last_Mon.day) + '-' + str(year) + ' ' +str(last_Sun_w) + '_' + str(last_Sun.day) + ' - PD_Stats_Report.xlsx')
    FN=str(year) + ' ' + str(last_Mon.year) + ' ' + str(last_Mon_w) + '_' + str(last_Mon.day) + '-' + str(year) + ' ' +str(last_Sun_w) + '_' + str(last_Sun.day) + ' - PD_Stats_Report.xlsx'
    # #檔名依產生日命名
# writer = pd.ExcelWriter(path + '\\' + str(year) + ' ' + str(last_month_w) + ' ' + str(xday.day) + ' - PD_Stats_Report.xlsx', engine='xlsxwriter', options={'remove_timezone': True})



## Write each dataframe to a different worksheet.

output.to_excel(writer, sheet_name='Raw_Data')
alert_count_per_shift_pt.to_excel(writer, sheet_name='Shift_stats')
alert_count_per_weekday_pt.to_excel(writer, sheet_name='Weekday_stats')
top_alerts.to_excel(writer, sheet_name='Top10')
alerts_escalation.to_excel(writer, sheet_name='# of escalation')
reassigned_to.to_excel(writer, sheet_name='Total # of alerts')
shift_comparison.to_excel(writer, sheet_name='Comparison')
Alert_Flood.to_excel(writer, sheet_name='Alert Flood')
top_host.to_excel(writer, sheet_name='Top_Host')
top_issue.to_excel(writer, sheet_name='Top_Issue')
# bytime.to_excel(writer, sheet_name='By_time')
# countalert.to_excel(writer, sheet_name='Alert_count')




# wb = openpyxl.load_workbook(path + '\\' + str(year) + ' ' + str(last_Mon_w) + '_' + str(last_Mon.day) +'-' +str(last_Sun_w) + '_' + str(last_Sun.day) + ' - PD_Stats_Report.xlsx')



###圖表參考資料 https://zhuanlan.zhihu.com/p/34286130
workbook  = writer.book
worksheet = writer.sheets['Shift_stats']

column1_chart = workbook.add_chart({'type': 'column'}) # 'type': 'column' 即為圖表類別為 bar chart
column1_chart.add_series({'name': 'PD alerts per day', # 'name': 顯示於 legend 的名稱
                      'categories': '=Shift_stats!$A$2:$A$8', # 'categories': X 軸值 
                      'values': '=Shift_stats!$E$2:$E$8', # 'values': Y 軸值
                      'data_labels':{'value':'true'}, # 'label的開關
                      
                        }) 
column1_chart.set_legend({'none': True,
                        })
column1_chart.options={'sort':False}

column1_chart.set_size({'width':600}) #'width':577,'height':287 圖表的大小
worksheet.insert_chart('A10', column1_chart) # 圖表擺放位置



worksheet = writer.sheets['Shift_stats']
column2_chart = workbook.add_chart({'type': 'column'}) # 'type': 'column' 即為圖表類別為 bar chart,寫在同一張圖這句只需一個
column2_chart.add_series({'name': 'Day', # 'name': 顯示於 legend 的名稱
                      'categories': '=Shift_stats!$A$2:$A$8', # 'categories': X 軸值 
                      'values': '=Shift_stats!$B$2:$B$8', # 'values': Y 軸值
                      'data_labels':{'value':'true'}, # 'label的開關
                      'fill':{'color':'#92D050'}
                      }) 

column2_chart.add_series({'name': 'Late', # 'name': 顯示於 legend 的名稱
                      'categories': '=Shift_stats!$A$2:$A$8', # 'categories': X 軸值 
                      'values': '=Shift_stats!$C$2:$C$8', # 'values': Y 軸值
                      'data_labels':{'value':'true'}, # 'label的開關
                      'fill':{'color':'#00B0F0'}
                        }) 

column2_chart.add_series({'name': 'Night', # 'name': 顯示於 legend 的名稱
                      'categories': '=Shift_stats!$A$2:$A$8', # 'categories': X 軸值 
                      'values': '=Shift_stats!$D$2:$D$8', # 'values': Y 軸值
                      'data_labels':{'value':'true'}, # 'label的開關
                      'fill':{'color':'#FFC000'}
                        }) 
column2_chart.set_legend({'position': 'top'})
column2_chart.set_size({'width':600}) #'width':577,'height':287
worksheet.insert_chart('K10', column2_chart) # 圖表擺放位置



# 找出總數
# wb = openpyxl.load_workbook(writer)
# sheet = wb.get_sheet_by_name('# of escalation')
# ttt = sheet.max_row
# tttt = str(ttt)
# print(tttt)
# worksheet = writer.sheets['# of escalation']
# test_chart = workbook.add_chart({'type': 'bar'})
# test_chart.add_series({'name': 'Escalation', # 'name': 顯示於 legend 的名稱
#                       'categories': '=# of escalation!$A$2:$A$:'+tttt, # 'categories': X 軸值 
#                       'values': '=# of escalation!$C$2:$C$:'+tttt, # 'values': Y 軸值
#                     #   'data_labels':{'value':'true'}, # 'label的開關
#                       'fill':{'color':'#FFC000'}
#                         }) 
# # test_chart.set_legend({'position': 'top'})
# # test_chart.set_size({'width':600}) #'width':577,'height':287
# worksheet.insert_chart('N2', test_chart) # 圖表擺放位置


worksheet = writer.sheets['Comparison']

pie_chart = workbook.add_chart({'type': 'pie'}) # 'type': 'column' 即為圖表類別為 bar chart
pie_chart.add_series({'name': 'Shift chart', # 'name': 顯示於 legend 的名稱
                      'categories': '=Comparison!$B$1:$D$1', # 'categories': X 軸值 
                      'values': '=Comparison!$B$2:$D$2', # 'values': Y 軸值
                      'data_labels':{'category':'true', 'percentage':'true'}, # 'label的開關
                      'points':[
                          {'fill':{'color':'#ED7D31'}},
                          {'fill':{'color':'#4472C4'}},
                          {'fill':{'color':'#A5A5A5'}},
                          ],
                        })


worksheet.insert_chart('A5', pie_chart) # 圖表擺放位置


## Close the Pandas Excel writer and output the Excel file.
writer.save()


print("Report export completed.")
print("Please check the folder ==>",path)
input('Press Any Key To Exit.') 



# In[ ]:




